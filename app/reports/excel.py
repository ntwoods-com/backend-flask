from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.utils.datetime import to_display_tz


def _auto_fit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
            except Exception:
                v = ""
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def _write_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append(r)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    _auto_fit(ws)


def build_workbook_bytes(
    *, report_type: str, from_s: str, to_s: str, timezone_display: str, data: dict[str, Any]
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    meta = wb.create_sheet("Meta")
    _write_table(
        meta,
        ["key", "value"],
        [
            ["type", report_type],
            ["from", from_s],
            ["to", to_s],
            ["generatedAt", to_display_tz(datetime.now(timezone.utc), timezone_display)],
        ],
    )

    if report_type == "summary":
        req = wb.create_sheet("Requirements")
        req_rows = [[r["status"], r["count"]] for r in data["summary"]["requirements"]["byStatus"]]
        _write_table(req, ["status", "count"], req_rows)

        cand = wb.create_sheet("Candidates")
        cand_rows = [[r["stage"], r["count"]] for r in data["summary"]["candidates"]["byStage"]]
        _write_table(cand, ["stage", "count"], cand_rows)
    else:
        ws = wb.create_sheet("Funnel")
        rows = [[r["stage"], r["count"]] for r in data["funnel"]["items"]]
        _write_table(ws, ["stage", "count"], rows)

    with BytesIO() as bio:
        wb.save(bio)
        return bio.getvalue()
